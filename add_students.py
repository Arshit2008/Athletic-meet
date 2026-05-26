import os
import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
import customtkinter as ctk

# Configure look and feel theme preferences
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class StudentRegistryApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Athletic Database Management Terminal")
        self.geometry("500x650")
        self.filename = "School_Reg.xlsx"
        self._initialize_excel_file()
        self._assemble_form_components()
        
    def _initialize_excel_file(self):
        """Creates the Excel master database file dynamically if missing"""
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Athletes_Master"
            headers = ["STU_ID", "NAME", "CLASS", "HOUSE", "DOB", "HEIGHT_CM", "WEIGHT_KG", "RUN_100M", "RUN_200M"]
            ws.append(headers)
            wb.save(self.filename)

    def _assemble_form_components(self):
        # Banner Header Frame
        banner = ctk.CTkLabel(self, text="ATHLETE MASTER REGISTRY", font=ctk.CTkFont(size=18, weight="bold"), text_color="#64B5F6", pady=15)
        banner.pack(fill=tk.X)
        
        # Scrollable Form Body Container to keep inputs clean on mobile
        form_frame = ctk.CTkScrollableFrame(self, width=450, height=480)
        form_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Student ID Entry Row
        ctk.CTkLabel(form_frame, text="Student ID (e.g. STU021):", font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10,2))
        self.ent_id = ctk.CTkEntry(form_frame, placeholder_text="Enter unique student ID ID...")
        self.ent_id.pack(fill=tk.X, padx=5)
        
        # Full Name Entry Row
        ctk.CTkLabel(form_frame, text="Competitor Full Name:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10,2))
        self.ent_name = ctk.CTkEntry(form_frame, placeholder_text="Enter full legal name...")
        self.ent_name.pack(fill=tk.X, padx=5)
        
        # Class & Section Layout Row Group Box Matrix Split
        class_section_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        class_section_frame.pack(fill=tk.X, pady=(10,2))
        
        # Left Side: Class Select Menu Dropdown 
        class_sub = ctk.CTkFrame(class_section_frame, fg_color="transparent")
        class_sub.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ctk.CTkLabel(class_sub, text="Class Grade:", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        classes_list = [f"{i}th" for i in range(1, 13)]
        self.drop_class = ctk.CTkComboBox(class_sub, values=classes_list, state="readonly")
        self.drop_class.set("12th")
        self.drop_class.pack(fill=tk.X, padx=(0,5))
        
        # Right Side: Section Select Menu Dropdown
        sec_sub = ctk.CTkFrame(class_section_frame, fg_color="transparent")
        sec_sub.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        ctk.CTkLabel(sec_sub, text="Section:", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        self.drop_sec = ctk.CTkComboBox(sec_sub, values=["A", "B", "C"], state="readonly")
        self.drop_sec.set("A")
        self.drop_sec.pack(fill=tk.X, padx=(5,0))
        
        # House Management Grid Selection Dropdown 
        ctk.CTkLabel(form_frame, text="Athletic House Team assignment:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10,2))
        self.drop_house = ctk.CTkComboBox(form_frame, values=["Red", "Blue", "Green", "Yellow"], state="readonly")
        self.drop_house.set("Red")
        self.drop_house.pack(fill=tk.X, padx=5)
        
        # Date of Birth String Entry Segment Input Block
        ctk.CTkLabel(form_frame, text="Date of Birth (DD-MM-YYYY format):", font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10,2))
        self.ent_dob = ctk.CTkEntry(form_frame, placeholder_text="e.g. 11-10-2008")
        self.ent_dob.pack(fill=tk.X, padx=5)
        
        # Biometric Dimension Metric Grid Layout Splits (Height/Weight values)
        biometrics_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        biometrics_frame.pack(fill=tk.X, pady=(10,2))
        
        # Height Field Value
        h_sub = ctk.CTkFrame(biometrics_frame, fg_color="transparent")
        h_sub.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ctk.CTkLabel(h_sub, text="Height (cm):", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        self.ent_height = ctk.CTkEntry(h_sub, placeholder_text="e.g. 175.5")
        self.ent_height.pack(fill=tk.X, padx=(0,5))
        
        # Weight Field Value
        w_sub = ctk.CTkFrame(biometrics_frame, fg_color="transparent")
        w_sub.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        ctk.CTkLabel(w_sub, text="Weight (kg):", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        self.ent_weight = ctk.CTkEntry(w_sub, placeholder_text="e.g. 68.0")
        self.ent_weight.pack(fill=tk.X, padx=(5,0))
        
        # Active Event Event Flag Checkboxes 
        ctk.CTkLabel(form_frame, text="Registered Track Events:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(15,2))
        self.check_100_var = ctk.StringVar(value="FALSE")
        self.check_100 = ctk.CTkCheckBox(form_frame, text="100m Sprint Event", variable=self.check_100_var, onvalue="TRUE", offvalue="FALSE")
        self.check_100.pack(anchor="w", padx=10, pady=5)
        
        self.check_200_var = ctk.StringVar(value="FALSE")
        self.check_200 = ctk.CTkCheckBox(form_frame, text="200m Sprint Event", variable=self.check_200_var, onvalue="TRUE", offvalue="FALSE")
        self.check_200.pack(anchor="w", padx=10, pady=5)
        
        # Master Submission Operations Execute Button Block
        save_btn = ctk.CTkButton(self, text="💾 COMMIT ATHLETE TO EXCEL MASTER", font=ctk.CTkFont(size=14, weight="bold"), height=45, fg_color="#2E7D32", hover_color="#1B5E20", command=self.save_record_to_excel)
        save_btn.pack(fill=tk.X, padx=25, pady=20)

    def save_record_to_excel(self):
        stu_id = self.ent_id.get().strip()
        name = self.ent_name.get().strip()
        cls_val = self.drop_class.get()
        sec_val = self.drop_sec.get()
        house = self.drop_house.get()
        dob = self.ent_dob.get().strip()
        h_str = self.ent_height.get().strip()
        w_str = self.ent_weight.get().strip()
        
        if not stu_id or not name or not dob or not h_str or not w_str:
            messagebox.showwarning("Validation Error", "All registration input fields must contain valid parameters!")
            return
            
        try:
            datetime.strptime(dob, "%d-%m-%Y")
        except ValueError:
            messagebox.showerror("Date Format Input Error", "Date of Birth entry details must match strict 'DD-MM-YYYY' format constraints!")
            return
            
        combined_class_section = f"{cls_val}-{sec_val}"
        
        try:
            wb = openpyxl.load_workbook(self.filename)
            sheet = wb.active
            
            # 1. FIXED DUPLICATE CHECKER: Safely scans existing entries
            is_duplicate = False
            if sheet.max_row >= 2:
                for r in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                    if r and r and str(r).strip() == stu_id:
                        is_duplicate = True
                        break
            
            if is_duplicate:
                messagebox.showerror("Key Integrity Crash", f"An athlete tracking index key matching record ID '{stu_id}' already exists!")
                wb.close()
                return
            
            # 2. FIXED APPEND MECHANISM: Force find the exact target row number
            target_row = sheet.max_row + 1
            row_data = [
                stu_id, name, combined_class_section, house, dob,
                float(h_str), float(w_str), self.check_100_var.get(), self.check_200_var.get()
            ]
            
            # Instead of using sheet.append(), we write cell-by-cell explicitly to block background mirroring
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=target_row, column=col_idx, value=value)
                
            wb.save(self.filename)
            wb.close() # Safely disconnect the file connection stream
            
            print(f"✅ SUCCESS: Saved student {stu_id} cleanly to Row {target_row}")
            messagebox.showinfo("Success", f"Registration complete! Athlete '{name}' recorded successfully.")
            
            # Reset UI inputs
            self.ent_id.delete(0, tk.END)
            self.ent_name.delete(0, tk.END)
            self.ent_dob.delete(0, tk.END)
            self.ent_height.delete(0, tk.END)
            self.ent_weight.delete(0, tk.END)
            self.check_100.deselect()
            self.check_200.deselect()
            
        except Exception as e:
            messagebox.showerror("Storage Fault", f"Disk structural execution locked:\n{str(e)}")
            
if __name__ == "__main__":
    app = StudentRegistryApp()
    app.mainloop()

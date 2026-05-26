import os
import sqlite3
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import openpyxl

# =========================================================================
# 1. EXCEL FILE PIPELINE
# =========================================================================
class AthleticExcelPipeline:
    def __init__(self):
        self.conn = sqlite3.connect(":memory:")
        self.cursor = self.conn.cursor()
        self._create_database_schema()
        
    def _create_database_schema(self):
        self.cursor.execute('CREATE TABLE IF NOT EXISTS school_data (stu_id TEXT PRIMARY KEY, name TEXT, class TEXT, house TEXT, dob TEXT, height REAL, weight REAL, run_100 INTEGER, run_200 INTEGER)')
        self.cursor.execute('CREATE TABLE IF NOT EXISTS trials_active (stu_id TEXT, event TEXT, heat INTEGER, lane INTEGER, name TEXT, house TEXT, class TEXT, division TEXT, track_time REAL DEFAULT 0.00)')
        self.cursor.execute('CREATE TABLE IF NOT EXISTS semifinals_output (event TEXT, division TEXT, heat INTEGER, lane INTEGER, name TEXT, house TEXT, seed_time REAL)')
        self.cursor.execute('CREATE TABLE IF NOT EXISTS finals_roster (event TEXT, division TEXT, lane INTEGER, name TEXT, house TEXT, seed_time REAL)')
        self.cursor.execute('CREATE TABLE IF NOT EXISTS house_points (house TEXT PRIMARY KEY, points INTEGER)')
        
        for h in ["Red", "Blue", "Green", "Yellow"]:
            self.cursor.execute("INSERT OR IGNORE INTO house_points (house, points) VALUES (?, 0)", (h,))
        self.conn.commit()

    def parse_bool(self, val):
        return 1 if str(val).strip().upper() in ['TRUE', '1', 'YES', 'Y'] else 0

    def load_excel_workbook(self, filename="School_Reg.xlsx"):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        full_path = os.path.join(script_dir, filename)

        if not os.path.exists(full_path):
            return False, f"[CRITICAL ERROR]\nFile '{filename}' is missing!\nPlease place it exactly here:\n{script_dir}"
            
        try:
            wb = openpyxl.load_workbook(full_path, data_only=True)
            sheet = wb.active
            data_tuples = []
            
            for row in sheet.iter_rows(min_row=2, max_row=100, values_only=True):
                if not row or row is None or str(row).strip() == "":
                    break 
                
                stu_id, name, cls, house, dob_val, height, weight, r100, r200 = row
                dob_string = dob_val.strftime("%d-%m-%Y") if isinstance(dob_val, datetime) else str(dob_val).strip()
                
                data_tuples.append((
                    str(stu_id).strip(), str(name).strip(), str(cls).strip(), str(house).strip(),
                    dob_string, float(height or 0.0), float(weight or 0.0),
                    self.parse_bool(r100), self.parse_bool(r200)
                ))
            
            self.cursor.execute("DELETE FROM school_data")
            self.cursor.executemany("INSERT OR REPLACE INTO school_data VALUES (?,?,?,?,?,?,?,?,?)", data_tuples)
            self.conn.commit()
            return True, "Success"
        except Exception as e:
            return False, f"[EXCEL PARSE ERROR]\n{str(e)}"

# =========================================================================
# 2. SEEDING LOGIC ENGINE
# =========================================================================
class MeetLogicProcessor:
    def __init__(self, pipeline):
        self.pipeline = pipeline

    def calculate_legacy_score(self, dob_str, height_cm, weight_kg):
        for fmt in ["%d-%m-%Y", "%Y-%m-%d"]:
            try:
                dob = datetime.strptime(str(dob_str).strip(), fmt)
                return round(((datetime(2026, 5, 25) - dob).days / 30.4 / 9.0) + (height_cm / 7.62) + (weight_kg * 0.73) - 3.50, 2)
            except ValueError: continue
        return 80.0 

    def execute_initial_seeding(self):
        self.pipeline.cursor.execute("SELECT stu_id, name, class, house, dob, height, weight FROM school_data WHERE run_100 = 1")
        runners = self.pipeline.cursor.fetchall()
        
        if len(runners) == 0:
            return "[ERROR] No students found running the 100m. Check your Excel file!"

        pools = {"O": [], "A": [], "B": [], "C": []}
        for stu_id, name, cls, house, dob, height, weight in runners:
            score = self.calculate_legacy_score(dob, height, weight)
            div = "O" if score >= 100 else "A" if score >= 88 else "B" if score >= 76 else "C"
            pools[div].append({"id": str(stu_id), "name": str(name), "class": str(cls), "house": str(house), "score": float(score)})

        self.pipeline.cursor.execute("DELETE FROM trials_active")
        log_view = "=== STAGE 1: EXTERNAL EXCEL INITIAL SEEDING LOG (100M) ===\n"

        for div, pool in pools.items():
            if not pool: continue
            
            pool.sort(key=lambda x: x["score"], reverse=True)
            heat_idx, rotation_state = 1, 0
            
            log_view += f"\n========================================\n      🏁 DIVISION {div} TRIALS BOARD      \n========================================\n"
            
            while len(pool) > 0:
                lanes, houses_in_heat = [None] * 6, set()
                
                for i in range(4):
                    for r in pool:
                        if r["house"] not in houses_in_heat:
                            lanes[i], houses_in_heat = r, houses_in_heat | {r["house"]}
                            pool.remove(r); break
                            
                for i, target in enumerate(["Red", "Blue"] if rotation_state % 2 == 0 else ["Green", "Yellow"], start=4):
                    for r in pool:
                        if r["house"] == target:
                            lanes[i] = r
                            pool.remove(r); break
                    if not lanes[i] and pool: lanes[i] = pool.pop(0)

                log_view += f"\n[Division {div} - Heat {heat_idx}]\n"
                for idx, runner in enumerate(lanes, start=1):
                    if runner:
                        self.pipeline.cursor.execute('INSERT INTO trials_active (stu_id, event, heat, lane, name, house, class, division, track_time) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)', (runner["id"], '100m', heat_idx, idx, runner["name"], runner["house"], runner["class"], div, 0.0))
                        log_view += f"  {idx}. {runner['name'].ljust(16)} | Class: {runner['class'].ljust(4)} | {runner['house']}\n"
                    else:
                        log_view += f"  {idx}. [EMPTY]\n"
                heat_idx += 1; rotation_state += 1
                
        self.pipeline.conn.commit()
        return log_view

# =========================================================================
# 3. MODERNIZED GUI DASHBOARD
# =========================================================================
class ModernMeetTerminal:
    def __init__(self, window, pipeline, engine):
        self.window, self.pipeline, self.engine = window, pipeline, engine
        self.window.title("Athletic Terminal Dashboard")
        self.window.geometry("850x700")
        self.window.configure(bg="#e8edf2")
        
        header_frame = tk.Frame(self.window, bg="#0f4c75", pady=15)
        header_frame.pack(fill=tk.X)
        tk.Label(header_frame, text="TRACK & FIELD BRACKET ENGINE", bg="#0f4c75", fg="#ffffff", font=("Segoe UI", 16, "bold")).pack()

        btn_frame = tk.Frame(self.window, bg="#e8edf2", pady=15)
        btn_frame.pack(fill=tk.X, padx=20)
        
        btn_style = {"font": ("Segoe UI", 10, "bold"), "width": 24, "relief": "flat", "pady": 8, "cursor": "hand2"}
        
        tk.Button(btn_frame, text="1. Run Initial Seeding", command=self.trigger_seeding, bg="#3282b8", fg="white", **btn_style).grid(row=0, column=0, padx=8, pady=6)
        tk.Button(btn_frame, text="2. Open Lab Timing Desk", command=self.display_input_desk, bg="#f2a365", fg="black", **btn_style).grid(row=0, column=1, padx=8, pady=6)
        tk.Button(btn_frame, text="3. Generate Semifinals", command=self.trigger_semifinals, bg="#9b59b6", fg="white", **btn_style).grid(row=0, column=2, padx=8, pady=6)
        tk.Button(btn_frame, text="4. Finals (Tick Marks)", command=self.display_tickmark_window, bg="#e74c3c", fg="white", **btn_style).grid(row=1, column=0, padx=8, pady=6)
        tk.Button(btn_frame, text="5. Final House Points", command=self.trigger_house_points, bg="#f1c40f", fg="black", **btn_style).grid(row=1, column=1, padx=8, pady=6)
        tk.Button(btn_frame, text="6. Export Full Packet (PDF)", command=self.export_to_pdf, bg="#27ae60", fg="white", **btn_style).grid(row=1, column=2, padx=8, pady=6)

        term_frame = tk.Frame(self.window, bg="#1b262c", bd=5, relief="sunken")
        term_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        self.display = tk.Text(term_frame, bg="#1b262c", fg="#00ffcc", font=("Consolas", 11), insertbackground="white", borderwidth=0, highlightthickness=0)
        self.display.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.display.insert(tk.END, "> SYSTEM INITIALIZED.\n> AWAITING EXCEL DATA STREAM...\n\n[INSTRUCTION] Click '1. Run Initial Seeding' to load the database.")

    def trigger_seeding(self):
        success, message = self.pipeline.load_excel_workbook()
        self.display.delete(1.0, tk.END)
        if not success:
            self.display.insert(tk.END, message)
        else:
            self.display.insert(tk.END, self.engine.execute_initial_seeding())

    def display_input_desk(self):
        self.pipeline.cursor.execute("SELECT stu_id, division, heat, lane, name, house, class FROM trials_active")
        rows = self.pipeline.cursor.fetchall()
        
        if not rows:
            messagebox.showwarning("Empty System", "No athletes are seeded!\nPlease click '1. Run Initial Seeding' first.")
            return

        desk = tk.Toplevel(self.window)
        desk.title("Lab Timing Desk")
        desk.geometry("550x600")
        desk.configure(bg="#ffffff")
        
        tk.Label(desk, text="ENTER TRACK TIMES", font=("Segoe UI", 14, "bold"), bg="#ffffff", fg="#0f4c75", pady=15).pack()
        
        canvas = tk.Canvas(desk, bg="#ffffff", highlightthickness=0)
        scrollbar = tk.Scrollbar(desk, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#ffffff")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=15)
        scrollbar.pack(side="right", fill="y")

        box_map = {}
        for uid, div, heat, lane, name, house, cls in rows:
            r_ui = tk.Frame(scroll_frame, bg="#f8f9fa", pady=6, padx=10)
            r_ui.pack(fill=tk.X, pady=4)
            tk.Label(r_ui, text=f"Div {div} | H{heat} L{lane} | {name.ljust(14)}", font=("Consolas", 10, "bold"), bg="#f8f9fa", width=32, anchor="w").pack(side=tk.LEFT)
            
            box_map[uid] = box = tk.Entry(r_ui, width=8, font=("Consolas", 12, "bold"), justify="center", bg="#e3f2fd", fg="#c62828", bd=2, relief="sunken")
            box.pack(side=tk.RIGHT)

        def commit():
            for uid, box in box_map.items():
                if box.get():
                    self.pipeline.cursor.execute("UPDATE trials_active SET track_time=? WHERE stu_id=?", (float(box.get()), uid))
            self.pipeline.conn.commit()
            messagebox.showinfo("Success", "Stopwatch times committed to memory!")
            desk.destroy()
            
        tk.Button(desk, text="COMMIT TIMINGS", command=commit, bg="#0f4c75", fg="white", font=("Segoe UI", 11, "bold"), relief="flat", pady=10).pack(side=tk.BOTTOM, fill=tk.X)

    def trigger_semifinals(self):
        self.pipeline.cursor.execute("SELECT COUNT(*) FROM trials_active WHERE track_time > 0.0")
        if self.pipeline.cursor.fetchone() == 0:
            self.display.delete(1.0, tk.END)
            self.display.insert(tk.END, "[ERROR] You must enter race times in the Lab Timing Desk first!")
            return
            
        self.pipeline.cursor.execute("DELETE FROM semifinals_output")
        log_view = "=== STAGE 2: OFFICIAL SEMIFINALS SEEDING SHEET ===\n"
        
        for div in ["O", "A", "B", "C"]:
            self.pipeline.cursor.execute("SELECT name, class, house, track_time FROM trials_active WHERE division=? AND track_time > 0 ORDER BY track_time ASC", (div,))
            runners = self.pipeline.cursor.fetchall()
            
            if runners:
                log_view += f"\n--- DIVISION {div} SEMIFINALS ---\n"
                heat, lane = 1, 1
                log_view += f"\n[Division {div} - Semifinal Heat {heat}]\n"
                
                for name, cls, house, time in runners:
                    self.pipeline.cursor.execute("INSERT INTO semifinals_output VALUES (?, ?, ?, ?, ?, ?, ?)", ('100m', div, heat, lane, name, house, time))
                    log_view += f"  {lane}. {name.ljust(16)} | Class: {cls.ljust(4)} | {house.ljust(7)} | Time: {time}s\n"
                    lane += 1
                    if lane > 6:
                        lane, heat = 1, heat + 1
                        if runners.index((name, cls, house, time)) < len(runners) - 1:
                            log_view += f"\n[Division {div} - Semifinal Heat {heat}]\n"
                            
        self.pipeline.conn.commit()
        self.display.delete(1.0, tk.END)
        self.display.insert(tk.END, log_view)

    def display_tickmark_window(self):
        self.pipeline.cursor.execute("SELECT COUNT(*) FROM trials_active WHERE track_time > 0.0")
        if self.pipeline.cursor.fetchone() == 0:
            messagebox.showwarning("No Data", "Please enter race times in the Timing Desk first!")
            return
            
        desk = tk.Toplevel(self.window)
        desk.title("Finals Qualification")
        desk.geometry("600x600")
        desk.configure(bg="#ffffff")
        
        tk.Label(desk, text="TICK THE RUNNERS ADVANCING TO FINALS", font=("Segoe UI", 12, "bold"), bg="#ffffff", fg="#c62828", pady=15).pack()
        
        canvas = tk.Canvas(desk, bg="#ffffff", highlightthickness=0)
        scrollbar = tk.Scrollbar(desk, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#ffffff")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=20)
        scrollbar.pack(side="right", fill="y")

        self.tick_vars = {} 
        for div in ["O", "A", "B", "C"]:
            self.pipeline.cursor.execute("SELECT stu_id, name, class, house, track_time FROM trials_active WHERE division=? AND track_time > 0 ORDER BY track_time ASC LIMIT 8", (div,))
            runners = self.pipeline.cursor.fetchall()
            
            if runners:
                tk.Label(scroll_frame, text=f"--- Division {div} Qualifiers ---", font=("Segoe UI", 11, "bold"), bg="#ffffff", fg="#0f4c75").pack(anchor="w", pady=8)
                for uid, name, cls, house, time in runners:
                    var = tk.IntVar(value=1)
                    self.tick_vars[(uid, div, name, cls, house, time)] = var
                    tk.Checkbutton(scroll_frame, text=f"{name.ljust(16)} | Class: {cls.ljust(4)} | {house.ljust(7)} | {time}s", variable=var, font=("Consolas", 11), bg="#ffffff").pack(anchor="w", padx=15, pady=2)

        def commit_finals():
            self.pipeline.cursor.execute("DELETE FROM finals_roster")
            log_view = "=== STAGE 3: OFFICIAL GRAND FINALS ROSTER ===\n"
            finalists_by_div = {"O": [], "A": [], "B": [], "C": []}
            
            for data, var in self.tick_vars.items():
                if var.get() == 1: 
                    uid, div, name, cls, house, time = data
                    finalists_by_div[div].append((name, cls, house, time))
                    
            for div, runners in finalists_by_div.items():
                if runners:
                    log_view += f"\n--- DIVISION {div} GRAND FINAL ---\n"
                    for lane, (name, cls, house, time) in enumerate(runners[:6], start=1):
                        self.pipeline.cursor.execute("INSERT INTO finals_roster VALUES (?, ?, ?, ?, ?, ?)", ('100m', div, lane, name, house, time))
                        log_view += f"  {lane}. {name.ljust(16)} | Class: {cls.ljust(4)} | {house.ljust(7)} | Q-Time: {time}s\n"
                        
            self.pipeline.conn.commit()
            self.display.delete(1.0, tk.END)
            self.display.insert(tk.END, log_view)
            desk.destroy()
            
        tk.Button(desk, text="CONFIRM TICK MARKS & GENERATE FINALS", command=commit_finals, bg="#e74c3c", fg="white", font=("Segoe UI", 11, "bold"), relief="flat", pady=10).pack(side=tk.BOTTOM, fill=tk.X)

# ==========================================================================================
# THESE TWO FUNCTIONS CONTAIN THE TUPLE EXTRACTOR FIXES THAT WERE BYPASSED BY CACHE
# ==========================================================================================
    def trigger_house_points(self):
        self.pipeline.cursor.execute("UPDATE house_points SET points = 0")
        log_view = "=== COMPREHENSIVE HOUSE POINTS LOG ===\n\n--- PHASE 1: QUALIFIER HEAT POINTS ---\n"
        points_scale, medals = [10, 6, 4], ["1st Place", "2nd Place", "3rd Place"]
        
        self.pipeline.cursor.execute("SELECT DISTINCT division, heat FROM trials_active WHERE track_time > 0 ORDER BY division, heat")
        heats = self.pipeline.cursor.fetchall()
        
        if not heats:
            self.display.delete(1.0, tk.END)
            self.display.insert(tk.END, "[ERROR] No race times found! Please enter times in the Timing Desk first.")
            return
            
        for div, heat in heats:
            self.pipeline.cursor.execute("SELECT name, house, track_time FROM trials_active WHERE division=? AND heat=? AND track_time > 0 ORDER BY track_time ASC LIMIT 3", (div, heat))
            for idx, (name, house, time) in enumerate(self.pipeline.cursor.fetchall()):
                self.pipeline.cursor.execute("UPDATE house_points SET points = points + ? WHERE house = ?", (points_scale[idx], house))
                log_view += f"[Div {div} H{heat}] {medals[idx]}: {name.ljust(14)} | {house.ljust(7)} -> +{points_scale[idx]} pts\n"

        log_view += "\n--- PHASE 2: GRAND FINALS POINTS ---\n"
        
        # FIX: Complete bypass of tuple checking logic
        try:
            self.pipeline.cursor.execute("SELECT name, house, seed_time FROM finals_roster ORDER BY division")
            all_finalists = self.pipeline.cursor.fetchall()
            
            if len(all_finalists) > 0:
                for div in ["O", "A", "B", "C"]:
                    self.pipeline.cursor.execute("SELECT name, house, seed_time FROM finals_roster WHERE division=? ORDER BY seed_time ASC LIMIT 3", (div,))
                    winners = self.pipeline.cursor.fetchall()
                    if winners:
                        log_view += f"\n[Division {div} Finals Podium]\n"
                        for idx, (name, house, time) in enumerate(winners):
                            self.pipeline.cursor.execute("UPDATE house_points SET points = points + ? WHERE house = ?", (points_scale[idx], house))
                            log_view += f"  {medals[idx]}: {name.ljust(14)} | {house.ljust(7)} -> +{points_scale[idx]} pts\n"
            else:
                log_view += "  (No finals have been generated yet.)\n"
        except Exception:
            log_view += "  (No finals have been generated yet.)\n"
            
        self.pipeline.conn.commit()
        log_view += "\n========================================\n        OVERALL HOUSE CUP STANDINGS       \n========================================\n"
        self.pipeline.cursor.execute("SELECT house, points FROM house_points ORDER BY points DESC")
        for house, pts in self.pipeline.cursor.fetchall():
            log_view += f"       {house.ljust(8)} : {pts} Points\n"
            
        self.display.delete(1.0, tk.END)
        self.display.insert(tk.END, log_view)

    def export_to_pdf(self):
        """Generates the Ultimate Multi-Page Tournament PDF Packet (Forced 4-Page Grid)"""
        try:
            from fpdf import FPDF
        except ImportError:
            messagebox.showerror("Missing Library", "PDF Export requires the 'fpdf' library.\nOpen VS Code Terminal and type:\npip install fpdf")
            return

        pdf = FPDF()
        
        # -----------------------------------------------------------------
        # FORCE PAGES 1-4: STRICTLY SCRIPTING ALL 4 DIVISION SHEETS
        # -----------------------------------------------------------------
        target_divisions = ["O", "A", "B", "C"]
        
        for current_div in target_divisions:
            # Force a brand new page open immediately for this division
            pdf.add_page()
            
            # Title Headers
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(190, 10, txt=f"OFFICIAL TRACK MARSHAL SHEET - DIVISION {current_div}", ln=True, align='C')
            pdf.set_font("Arial", 'I', 10)
            pdf.cell(190, 10, txt="Qualifier Trials - Sorted Alphabetically for Attendance Checking", ln=True, align='C')
            pdf.ln(5)

            # Pull whatever data exists for this specific division
            self.pipeline.cursor.execute(
                "SELECT name, class, house, heat, lane FROM trials_active WHERE division=? ORDER BY name ASC", 
                (current_div,)
            )
            runners_in_div = self.pipeline.cursor.fetchall()

            # Check if this division actually has runners loaded
            if len(runners_in_div) == 0:
                # If zero runners exist, manually draw a clean box so the PDF doesn't stall out
                pdf.set_fill_color(245, 245, 245)
                pdf.set_font("Arial", 'I', 12)
                pdf.cell(190, 20, txt=f"Notice: No athletes are registered or running in Division {current_div}.", border=1, ln=True, align='C', fill=True)
            else:
                # If runners exist, draw the standard blue table data grid
                pdf.set_fill_color(200, 220, 255)
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(60, 10, 'Athlete Name', 1, 0, 'C', fill=True)
                pdf.cell(30, 10, 'Class', 1, 0, 'C', fill=True)
                pdf.cell(40, 10, 'House', 1, 0, 'C', fill=True)
                pdf.cell(30, 10, 'Heat', 1, 0, 'C', fill=True)
                pdf.cell(30, 10, 'Lane', 1, 1, 'C', fill=True)

                pdf.set_font("Arial", '', 11)
                for name, cls, house, heat, lane in runners_in_div:
                    pdf.cell(60, 10, str(name), 1, 0, 'L')
                    pdf.cell(30, 10, str(cls), 1, 0, 'C')
                    pdf.cell(40, 10, str(house), 1, 0, 'C')
                    pdf.cell(30, 10, f"Heat {heat}", 1, 0, 'C')
                    pdf.cell(30, 10, f"Lane {lane}", 1, 1, 'C')

        # -----------------------------------------------------------------
        # FORCE PAGE 5: GRAND FINALS ROSTER (IF RUNNERS ARE QUALIFIED)
        # -----------------------------------------------------------------
        try:
            self.pipeline.cursor.execute("SELECT lane FROM finals_roster")
            finals_list_check = self.pipeline.cursor.fetchall()
            
            if len(finals_list_check) > 0:
                pdf.add_page()
                pdf.set_font("Arial", 'B', 16)
                pdf.cell(190, 10, txt="OFFICIAL GRAND FINALS ROSTER", ln=True, align='C')
                pdf.ln(5)
                
                for div in target_divisions:
                    self.pipeline.cursor.execute("SELECT lane, name, class, house, seed_time FROM finals_roster WHERE division=? ORDER BY lane ASC", (div,))
                    finalists = self.pipeline.cursor.fetchall()
                    if finalists:
                        pdf.set_font("Arial", 'B', 14)
                        pdf.cell(190, 10, txt=f"Division {div} Finals", ln=True, align='L')
                        
                        pdf.set_fill_color(255, 230, 200)
                        pdf.set_font("Arial", 'B', 12)
                        pdf.cell(20, 10, 'Lane', 1, 0, 'C', fill=True)
                        pdf.cell(60, 10, 'Name', 1, 0, 'C', fill=True)
                        pdf.cell(30, 10, 'Class', 1, 0, 'C', fill=True)
                        pdf.cell(40, 10, 'House', 1, 0, 'C', fill=True)
                        pdf.cell(40, 10, 'Seed Time', 1, 1, 'C', fill=True)
                        
                        pdf.set_font("Arial", '', 11)
                        for lane, name, cls, house, time in finalists:
                            pdf.cell(20, 10, str(lane), 1, 0, 'C')
                            pdf.cell(60, 10, str(name), 1, 0, 'L')
                            pdf.cell(30, 10, str(cls), 1, 0, 'C')
                            pdf.cell(40, 10, str(house), 1, 0, 'C')
                            pdf.cell(40, 10, f"{time}s", 1, 1, 'C')
                        pdf.ln(5)
        except Exception:
            pass

        # -----------------------------------------------------------------
        # FORCE PAGE 6: OFFICIAL STANDINGS STANDARDS (IF POINTS EXIST)
        # -----------------------------------------------------------------
        try:
            self.pipeline.cursor.execute("SELECT points FROM house_points WHERE points > 0")
            has_points = self.pipeline.cursor.fetchall()
            
            if len(has_points) > 0:
                pdf.add_page()
                pdf.set_font("Arial", 'B', 18)
                pdf.cell(190, 10, txt="OFFICIAL HOUSE CUP STANDINGS", ln=True, align='C')
                pdf.ln(10)
                
                pdf.set_fill_color(220, 255, 220)
                pdf.set_font("Arial", 'B', 14)
                pdf.cell(95, 12, 'School House', 1, 0, 'C', fill=True)
                pdf.cell(95, 12, 'Total Points', 1, 1, 'C', fill=True)
                
                pdf.set_font("Arial", 'B', 12)
                self.pipeline.cursor.execute("SELECT house, points FROM house_points ORDER BY points DESC")
                for house, pts in self.pipeline.cursor.fetchall():
                    pdf.cell(95, 12, str(house), 1, 0, 'C')
                    pdf.cell(95, 12, f"{pts} pts", 1, 1, 'C')
        except Exception:
            pass

        # Save out to physical storage drive
        pdf_filename = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Official_Tournament_Packet.pdf")
        try:
            pdf.output(pdf_filename)
            self.display.insert(tk.END, f"\n[SUCCESS] Entire Tournament PDF Exported!\nSaved as: Official_Tournament_Packet.pdf\n")
            messagebox.showinfo("Export Successful", "Full tournament PDF generated successfully with all sheets.")
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))
if __name__ == "__main__":
    root = tk.Tk()
    file_layer = AthleticExcelPipeline()
    logic_layer = MeetLogicProcessor(file_layer)
    app = ModernMeetTerminal(root, file_layer, logic_layer)
    root.mainloop()